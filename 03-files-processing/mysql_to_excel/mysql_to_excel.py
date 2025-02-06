# -*- coding: utf-8 -*-
"""
执行多个sql，并将数据分别写入同个excel文件的不同页签，sql及数据库配置方式见 sql.xlsx
"""

import os
import pandas as pd
import pymysql
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from time import strftime

def connect_to_database(df_conf):
    try:
        return pymysql.connect(
            host=df_conf.iloc[0, 2].split("/")[0].split(":")[0],
            port=int(df_conf.iloc[0, 2].split("/")[0].split(":")[1]),
            user=df_conf.iloc[0, 0],
            password=df_conf.iloc[0, 1],
            database=df_conf.iloc[0, 2].split("/")[1],
            charset='utf8'
        )
    except Exception as e:
        print(f'连接mysql数据库出错: ', e)
        return None

def execute_sql_and_write_to_excel(df_sql, connection, out_path):
    cursor = connection.cursor(cursor=pymysql.cursors.DictCursor)
    
    book = load_workbook(out_path)

    for i, row in df_sql.iterrows():
        sql = row['sql']
        sheet_name = row['sheetName']
        try:
            cursor.execute(sql)
            title = [j[0] for j in cursor.description]
            new_df = pd.DataFrame(cursor.fetchall())
            
            # 将 DataFrame 中的 None 值替换为空字符串
            new_df = new_df.fillna('')

            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                next_col = sheet.max_column + 1
                for idx, col_name in enumerate(title):
                    sheet.cell(row=1, column=next_col + idx, value=col_name)
                for row_idx, data in enumerate(new_df.values):
                    for col_idx, value in enumerate(data):
                        sheet.cell(row=row_idx + 2, column=next_col + col_idx, value=value)
            else:
                sheet = book.create_sheet(title=sheet_name)
                for r in dataframe_to_rows(new_df, index=False, header=True):
                    sheet.append(r)
            
            print(f'已完成 {sheet_name} 的输出。')
        except Exception as e:
            print(f'{sheet_name} 执行结果为空或出错: {e}')
    
    book.save(out_path)
    cursor.close()

def process_file(df_xlsx, out_path):
    try:
        df_sql = pd.read_excel(f'./{df_xlsx}', sheet_name=0)
        df_conf = pd.read_excel(f'./{df_xlsx}', sheet_name=1)

        connection = connect_to_database(df_conf)
        if connection:
            execute_sql_and_write_to_excel(df_sql, connection, out_path)
            connection.close()
        else:
            print('数据库连接失败，跳过该文件。')
    except Exception as e:
        print(f'执行文件有误!! 跳过本次: {e}')

if __name__ == "__main__":
    if not os.path.exists('./out'):
        os.makedirs('./out')

    systime = strftime("%Y%m%d_%H%M%S")
    out_path = f'./out/output_{systime}.xlsx'
    dataset = pd.DataFrame([[systime, "TEST"]], columns=["文件生成时间", "测试列"])
    with pd.ExcelWriter(out_path) as writer:
        dataset.to_excel(writer, sheet_name='测试页', header=True, index=False)

    print('输出excel小程序开始执行。。。')
    df_xlsxs = input("请输入读取文件名(如sql.xlsx, 可用逗号分隔, 如 sql,stcmm,mpay )：")
    files = df_xlsxs.split(",")

    print(f'共{len(files)}个文件需要执行')
    for i, file in enumerate(files):
        df_xlsx = f'{file}.xlsx'
        print(f'\n当前执行第{i+1}个文件：{df_xlsx}')
        process_file(df_xlsx, out_path)

    input(f'\n执行完毕！！结果已输出至{out_path} 回车退出。。。')